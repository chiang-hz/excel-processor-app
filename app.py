# app.py

import streamlit as st
import pandas as pd
from fpdf import FPDF
import openpyxl  # 我們需要 openpyxl 來讀取頁面設定
import os
from io import BytesIO

# --- 1. 新增：設定提取函式 ---

def extract_page_setup_from_excel(uploaded_file) -> dict:
    """
    使用 openpyxl 從上傳的 Excel 檔案中提取頁面排版設定。

    Args:
        uploaded_file: Streamlit 的 UploadedFile 物件。

    Returns:
        dict: 包含提取設定的字典，如果失敗則返回空字典。
    """
    settings = {}
    try:
        # 將 UploadedFile 物件傳遞給 openpyxl
        # uploaded_file 可以像檔案一樣被讀取
        workbook = openpyxl.load_workbook(uploaded_file)
        # 以第一個活動的工作表為準來讀取設定
        worksheet = workbook.active
        ps = worksheet.page_setup

        # --- 單位與代碼對應表 ---
        # openpyxl 的邊距單位是英寸，1 英寸 = 2.54 公分
        INCH_TO_CM = 2.54
        # openpyxl 的紙張大小是字串代碼
        PAPER_SIZE_MAP_REVERSE = {'1': 'Letter', '8': 'A3', '9': 'A4'}
        ORIENTATION_MAP_REVERSE = {'portrait': '直向', 'landscape': '橫向'}
        
        # --- 提取與轉換 ---
        settings['紙張大小'] = PAPER_SIZE_MAP_REVERSE.get(ps.paperSize, 'A4')
        settings['頁面方向'] = ORIENTATION_MAP_REVERSE.get(ps.orientation, '直向')
        
        # 邊距 (英寸 -> 公分)，並四捨五入到小數點後一位
        settings['上邊距'] = round(ps.top * INCH_TO_CM, 1) if ps.top else 1.9
        settings['下邊距'] = round(ps.bottom * INCH_TO_CM, 1) if ps.bottom else 1.5
        settings['左邊距'] = round(ps.left * INCH_TO_CM, 1) if ps.left else 1.2
        settings['右邊距'] = round(ps.right * INCH_TO_CM, 1) if ps.right else 1.2
        
        # 提取頁尾 (以中間的頁尾為準)
        # openpyxl 的頁尾代碼與 pywin32 類似: &P=頁碼, &N=總頁數
        if worksheet.footer.center.text:
            settings['頁尾內容'] = worksheet.footer.center.text
        else:
            settings['頁尾內容'] = "第 &P 頁 / 共 &N 頁"
            
        return settings

    except Exception as e:
        # 如果讀取失敗 (例如檔案格式問題)，返回空字典並在控制台打印錯誤
        print(f"無法讀取 Excel 設定: {e}")
        return {}


# --- 2. 核心 PDF 處理函式 (與之前相同) ---
class PDF(FPDF):
    def __init__(self, orientation='P', unit='mm', format='A4', footer_text=""):
        super().__init__(orientation, unit, format)
        self.footer_text = footer_text
        self.set_auto_page_break(auto=True, margin=15)

    def footer(self):
        self.set_y(-15)
        self.set_font('NotoSans', '', 8)
        # 替換頁碼代碼
        final_footer_text = self.footer_text.replace('&P', str(self.page_no())).replace('&N', '{nb}')
        self.cell(0, 10, final_footer_text, 0, 0, 'C')

def process_excel_to_pdf_cross_platform(uploaded_file, options: dict) -> bytes:
    try:
        # 重設讀取指標，因為 uploaded_file 可能已被 extract_page_setup_from_excel 讀取過
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file)
        sheets_data = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}

        orientation_map = {'直向': 'P', '橫向': 'L'}
        pdf = PDF(
            orientation=orientation_map.get(options['頁面方向'], 'P'),
            unit='mm',
            format=options['紙張大小'],
            footer_text=options['頁尾內容']
        )
        
        font_path = 'NotoSansTC-Regular.ttf'
        if not os.path.exists(font_path):
            st.error(f"錯誤：找不到字體檔案 '{font_path}'。請確保它與 app.py 放在一起。")
            st.error("可從 Google Fonts 下載：https://fonts.google.com/noto/specimen/Noto+Sans+TC")
            return None
            
        pdf.add_font('NotoSans', '', font_path, uni=True)
        pdf.set_font('NotoSans', '', 10)

        pdf.set_top_margin(options['上邊距'] * 10)
        pdf.b_margin = options['下邊距'] * 10
        pdf.set_left_margin(options['左邊距'] * 10)
        pdf.set_right_margin(options['右邊距'] * 10)

        for sheet_name, df in sheets_data.items():
            pdf.add_page()
            pdf.set_font('NotoSans', '', 14)
            pdf.cell(0, 10, sheet_name, 0, 1, 'L')
            
            headers = [str(col) for col in df.columns]
            if not headers: continue # 跳過空的工作表

            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            col_width = page_width / len(headers)
            
            pdf.set_fill_color(230, 230, 230)
            pdf.set_font('NotoSans', '', 10)
            for header in headers:
                pdf.cell(col_width, 8, header, 1, 0, 'C', fill=True)
            pdf.ln()

            pdf.set_font('NotoSans', '', 9)
            for index, row in df.iterrows():
                str_row = [str(item) if pd.notna(item) else "" for item in row]
                # 使用 get_string_width 來處理換行，這裡簡化為統一高度
                # 較複雜的表格需要更精密的計算
                max_y = pdf.get_y()
                for i, item in enumerate(str_row):
                    x = pdf.l_margin + i * col_width
                    pdf.set_xy(x, pdf.get_y())
                    pdf.multi_cell(col_width, 8, item, border=1, align='L')
                    if pdf.get_y() > max_y:
                        max_y = pdf.get_y()
                pdf.set_y(max_y)

        pdf.alias_nb_pages()
        return pdf.output(dest='S').encode('latin-1')
    except Exception as e:
        st.error(f"PDF 轉換失敗: {e}")
        return None

# --- 3. Streamlit 應用程式介面 (UI) ---

st.set_page_config(page_title="智慧 Excel to PDF 轉換器", layout="centered")

st.title("📄 智慧 Excel to PDF 轉換器")
st.info("上傳 Excel 後，系統將自動讀取其頁面設定。您可以在側邊欄微調後再產生 PDF。")

# 初始化 session_state
if 'extracted_settings' not in st.session_state:
    st.session_state.extracted_settings = {}
if 'last_file_id' not in st.session_state:
    st.session_state.last_file_id = None

# --- 4. 側邊欄設定介面 ---
with st.sidebar:
    st.header("⚙️ 排版設定")
    
    uploaded_file = st.file_uploader(
        "上傳 Excel 檔案", 
        type=['xlsx', 'xls'],
        help="支援 .xlsx 和 .xls 格式的檔案。"
    )

    # --- 自動提取設定的邏輯 ---
    if uploaded_file is not None:
        # 檢查是否是新的檔案，避免重複執行
        if uploaded_file.file_id != st.session_state.get('last_file_id'):
            with st.spinner("正在讀取 Excel 設定..."):
                st.session_state.extracted_settings = extract_page_setup_from_excel(uploaded_file)
                st.session_state.last_file_id = uploaded_file.file_id
            
            if st.session_state.extracted_settings:
                st.success("已成功讀取並套用 Excel 排版設定！")
            else:
                st.warning("無法讀取此檔案的排版設定，將使用預設值。")

    # 使用 .get() 從 session_state 安全地讀取設定，如果沒有則使用預設值
    defaults = st.session_state.extracted_settings
    
    with st.form(key="settings_form"):
        st.subheader("頁面配置")
        
        # 動態設定 index
        paper_options = ['A4', 'A3', 'Letter']
        paper_index = paper_options.index(defaults.get('紙張大小', 'A4'))
        paper_size = st.selectbox("紙張大小", options=paper_options, index=paper_index)
        
        orient_options = ['直向', '橫向']
        orient_index = orient_options.index(defaults.get('頁面方向', '直向'))
        page_orientation = st.radio("頁面方向", options=orient_options, index=orient_index, horizontal=True)

        st.subheader("頁面邊距 (公分)")
        col1, col2 = st.columns(2)
        with col1:
            top_margin = st.number_input("上邊距", value=defaults.get('上邊距', 1.9), min_value=0.5, step=0.1, format="%.1f")
            bottom_margin = st.number_input("下邊距", value=defaults.get('下邊距', 1.5), min_value=0.5, step=0.1, format="%.1f")
        with col2:
            left_margin = st.number_input("左邊距", value=defaults.get('左邊距', 1.2), min_value=0.5, step=0.1, format="%.1f")
            right_margin = st.number_input("右邊距", value=defaults.get('右邊距', 1.2), min_value=0.5, step=0.1, format="%.1f")
        
        st.subheader("頁尾設定")
        footer_text = st.text_input("頁尾內容", value=defaults.get('頁尾內容', '第 &P 頁 / 共 &N 頁'), help="使用 &P 代表頁碼, &N 代表總頁數。")
        
        submit_button = st.form_submit_button(label="🚀 產生 PDF")

# --- 5. 主程式邏輯 ---
if submit_button:
    if uploaded_file is not None:
        with st.spinner("正在轉換為 PDF，請稍候..."):
            format_options = {
                '紙張大小': paper_size,
                '頁面方向': page_orientation,
                '上邊距': top_margin,
                '下邊距': bottom_margin,
                '左邊距': left_margin,
                '右邊距': right_margin,
                '頁尾內容': footer_text,
            }

            # 確保檔案指標在開頭，以便 PDF 轉換函式能讀取
            uploaded_file.seek(0)
            pdf_bytes = process_excel_to_pdf_cross_platform(uploaded_file, format_options)

            if pdf_bytes:
                st.success("PDF 產生成功！")
                file_name = f"{os.path.splitext(uploaded_file.name)[0]}_converted.pdf"
                st.download_button(
                    label="📥 下載 PDF",
                    data=pdf_bytes,
                    file_name=file_name,
                    mime="application/pdf"
                )
    else:
        st.warning("請先上傳一個 Excel 檔案！")